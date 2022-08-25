import tkinter
import random
from tkinter.constants import CENTER
from tkinter import messagebox

#importing openpyxl and creating a sheet via work book
import openpyxl
loc="final.xlsx"
wb=openpyxl.load_workbook(loc)
sheet=wb.active
row=sheet.max_row
col=sheet.max_column

def start_game(root):
    global chosen_word, word_given
    global turns2
    global guess, top
    global text_guess, text_turns
    global e, b

    #deleting the result row
    global row
    global col
    improvise=sheet.cell(row=1,column=col)
    if improvise.value=="Result":
        improvise.value=None
        sheet.delete_cols(col)
        row=sheet.max_row
        col=sheet.max_column
        change_1=sheet.cell(row=1,column=col+1)
        change_1.value="Hangman"
        change_2=sheet.cell(row=2,column=col+1)
        change_2.value="Singleplayer"
    else:
        row=sheet.max_row
        col=sheet.max_column

        change_1=sheet.cell(row=1,column=col+1)
        change_1.value="Hangman"
        change_2=sheet.cell(row=2,column=col+1)
        change_2.value="Singleplayer"
        

    words = ['apple', 'banana', 'grape', 'kiwi', 'mango', 'orange', 'potato', 'tomato', 'onion', 'lion','tiger', 'leopard']
    word_given = random.choice(words)
    chosen_word = list(word_given)
    turns2 = 6

    top = tkinter.Toplevel(root)
    top.title("Hangman")
    top.minsize(width = 400, height = 400)

    guess = []
    for i in range(len(chosen_word)):
        guess.append('_')

    b = "  "
    text_guess = tkinter.Label(top, text = b.join(guess), font=("Arial", 28))
    text_guess.place(relx=0.4, rely=0.15, anchor = CENTER)
    text_turns = tkinter.Label(top, text = "turns left - {}".format(turns2), font=("Arial", 28))
    text_turns.place(relx=0.4, rely=0.35, anchor = CENTER)

    get_letter = tkinter.Label(top, text = "enter a letter: - ", font=("Arial", 28)).place(relx=0.4, rely=0.55, anchor = CENTER)
    e = tkinter.Entry(top, width = 10, font=("Arial", 10))
    e.place(relx=0.8, rely=0.55, anchor = CENTER)

    ok_button = tkinter.Button(top, text = "enter", font=("Arial", 20),  command = button_click).place(relx=0.4, rely=0.75, anchor = CENTER)

    top.mainloop()

def button_click():
    global turns2
    ch = e.get().lower()
    e.delete(0, 'end')
    found = 0
    indices = [i for i, x in enumerate(chosen_word) if x == ch]
    if len(indices) != 0:
        found = 1
    for i in indices:
        guess[i] = ch
    if found == 0:
        turns2 = turns2 - 1
    text_guess.config(text = b.join(guess))
    text_turns.config(text = "turns left - {}".format(turns2))

    if turns2 == 0:
        messagebox.showinfo('LOST', 'Sorry, you lost, the word was {}'.format(word_given))
        top.destroy()
    
    if guess == chosen_word:
        messagebox.showinfo('WON', 'HOORAY YOU WON!!!!!')
        top.destroy()

    if guess == chosen_word or turns2 == 0:
        reply=messagebox.askquestion('Game Continuation Confirmation','Do you want to play more?')
        if reply=='yes':
            messagebox.showinfo("Continue",'Click on a GameIcon')
        else:
            messagebox.showinfo("EXIT",'Click on EXIT')

    global row
    global column
    c1=sheet.cell(row=3,column=col+1)
    c2=sheet.cell(row=4,column=col+1)
    
    #appending scores
    if turns2 == 0:
        c1.value=0
        c2.value=10
    
    if guess == chosen_word:
        c1.value=10
        c2.value=0


    wb.save('final.xlsx')



